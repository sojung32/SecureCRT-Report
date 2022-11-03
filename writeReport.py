# $language = "python3"
# $interface = "1.0"

# This code must be run on SecureCRT!!!
#
# Descript
## Import Cisco devices information from Excel
## Connect to server using SSH2 protocol
## Excute a specific Commands
## Then Write down Prompt to an Excel and Text File
#
# Required 
## SecureCRT version 9.x.x or later
## Python3 (3.9.13 recommended)
## pandas and openpyxl Required
## The Excel File containing host port username password

import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from datetime import datetime


### Open Excel File
def getExcel():
	return crt.Dialog.FileOpenDialog(title="Open File", filter="Excel Files (*.xlsx)|*.xlsx||")


### Read Excel File and Get Information
def getInformation(file):
	try:
		infoExcel = pd.read_excel(file, engine='openpyxl')
		infoFrame = pd.DataFrame(infoExcel)

		return infoFrame

	except Exception as err:
		crt.Dialog.MessageBox(str(err))


### Try Connect using SSH2 protocol
def connectSSH2(host, port, username, password):
	# Session Connection
	try:
		crt.Session.Connect("/SSH2 %s /P %s /L %s /password %s" % (host, port, username, password))

	# Error Occurred
	except ScriptError:
		err = crt.GetLastErrorMessage()
		crt.Dialog.MessageBox(err)


### Excute Commands
### Get Command Results
### Set Data to Excel and Save Text File
def excuteCommand(sheet, index):
	# Command Array
	cmd = ['show version | include Model', 
			'sh run | include hostname|Hostname', 
			'show version | include Version',
			'show version | include Uptime|uptime',
			'show processes cpu | include CPU',
			'show processes memory | include Processor',
			'dir | include total',
			'show env all']

	logCmd = ['sh run', 'show logging']
	txtName = ['_config.txt', '_log.txt'] # text file save name

	crt.Screen.Synchronous = True
	crt.Screen.IgnoreEscape = True
	dataRow = 5

	# Excute Commands for Excel
	for c in cmd:
		while True : 
			if not crt.Screen.WaitForCursor(1):
				break

		# Get Start Row Line
		startRow = crt.Screen.CurrentRow
		prompt = crt.Screen.Get(startRow, 0, startRow, crt.Screen.CurrentColumn)
		prompt = prompt.strip()

		# Excute Command
		crt.Screen.Send(c + "\r")
		crt.Screen.WaitForString(prompt)

		screenrow = crt.Screen.CurrentRow
		bContinue = True
		
		# Set Result Data to Excel
		for i in range(1, screenrow) :
			result = crt.Screen.Get(i, 0, i, 1000)

			if result.find("#") < 0:
				# model
				if c.find("Model"):
					regex = re.compile("Model Number|Model number")
					if regex.search(result):
						cell = sheet.cell(dataRow+index, 1, result.split(':')[-1].strip())
						setCellStyleDefault(cell)
				# hostname
				if c.find("Hostname"):
					regex = re.compile("hostname|Hostname")
					if regex.search(result):
						hostname = result.split()[-1].strip()
						cell = sheet.cell(dataRow+index, 2, hostname)
						setCellStyleDefault(cell)
				# version
				if c.find("Version") and bContinue:
					if result.find("Version") > -1:
						version = result.split('Version')[-1].split(",")[0].strip()
						cell = sheet.cell(dataRow+index, 3, version)
						setCellStyleDefault(cell)
						bContinue = False
				# uptime
				if c.find("Uptime"):
					regex = re.compile("Uptime|uptime")
					if regex.search(result):
						uptime = result.split("is")[-1]
						if uptime.find("minutes"):
							uptime = uptime[0:uptime.rfind(',')]
						cell = sheet.cell(dataRow+index, 4, uptime)
						setCellStyleDefault(cell)
				# cpu
				if c.find("CPU") and bContinue:
					if result.find("CPU") > -1:
						cell = sheet.cell(dataRow+index, 5, result.split('five minutes:')[-1].strip())
						setCellStyleDefault(cell)
						bContinue = False
				# memory
				if c.find("memory") and bContinue:
					if result.find("Processor") > -1:
						total = result.split("Total:")[-1].split("Used:")[0].strip()
						used = result.split("Used:")[-1].split("Free:")[0].strip()
						cell = sheet.cell(dataRow+index, 6, total)
						setCellStyleDefault(cell)
						cell = sheet.cell(dataRow+index, 7, used)
						setCellStyleDefault(cell)
						bContinue = False
				# flash
				if c.find("total"):
					if result.find("total") > -1:
						flash = result.split("total")
						cell = sheet.cell(dataRow+index, 8, flash[0].split()[0].strip())
						setCellStyleDefault(cell)
						cell = sheet.cell(dataRow+index, 9, re.sub("[\\(|\\)]", "", flash[1]).split()[0].strip())
						setCellStyleDefault(cell)
				# env(temp, power, fan)
				if c.find("env"):
					#temp
					if result.find("Temperature Value") > -1:
						cell = sheet.cell(dataRow+index, 10, re.sub("[^0-9]", "", result).strip())
						setCellStyleDefault(cell)
					#power
					if result.find("Built-in") > -1:
						cell = sheet.cell(dataRow+index, 11, result.split()[-1].strip())
						setCellStyleDefault(cell)
					# fan
					regex = re.compile("Fan|FAN")
					if regex.search(result):
						cell = sheet.cell(dataRow+index, 12, result.split("is")[-1].strip())
						setCellStyleDefault(cell)

		# Clear Screen
		crt.Screen.Clear()
		crt.Screen.Send("\r")
		crt.Screen.WaitForCursor(1)

	# Excute Commands for Text File
	for (idx, c) in enumerate(logCmd):
		while True : 
			if not crt.Screen.WaitForCursor(1):
				break

		# Get Start Row Line
		startRow = crt.Screen.CurrentRow
		prompt = crt.Screen.Get(startRow, 0, startRow, crt.Screen.CurrentColumn)
		prompt = prompt.strip()

		# Excute Command
		crt.Screen.Send(c + "\r")
		crt.Screen.WaitForString("\r", 1)
		crt.Screen.WaitForString("\n", 1)

		# Save Result Data to Text Files
		result = re.sub("\n+", "", crt.Screen.ReadString(prompt)).strip()
		txtFile = open(savePath + hostname + txtName[idx], "w")
		txtFile.write(result)
		txtFile.close()

	crt.Screen.Clear()
	crt.Screen.Synchronous = False


### Prepare Excel Writing - setting titles
def prepareExcel(sheet):

	global alignCenter, fontBold, borderThin, fillColor

	# cell styles
	alignCenter = Alignment(horizontal="center", vertical="center")
	alignRight = Alignment(horizontal="right", vertical="center")
	fontBold = Font(bold=True)
	borderThin = Border(left=Side(style="thin"), 
						right=Side(style="thin"),
						top=Side(style="thin"),
						bottom=Side(style="thin"))
	fillColor = PatternFill(fgColor="B7DEE8", fill_type="solid")

	# cell height, width setting
	sheet.row_dimensions[1].height = 25
	sheet.column_dimensions['A'].width = 20
	sheet.column_dimensions['B'].width = 15
	sheet.column_dimensions['C'].width = 15
	sheet.column_dimensions['D'].width = 15
	sheet.column_dimensions['E'].width = 10
	sheet.column_dimensions['F'].width = 12
	sheet.column_dimensions['G'].width = 12
	sheet.column_dimensions['H'].width = 12
	sheet.column_dimensions['I'].width = 12
	sheet.column_dimensions['J'].width = 10
	sheet.column_dimensions['K'].width = 10
	sheet.column_dimensions['L'].width = 10

	# title
	sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
	cellTitle = sheet.cell(row=1, column=1, value='Report')
	cellTitle.font = Font(size=15, bold=True)
	cellTitle.alignment = alignCenter

	# Date
	sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
	cellDate = sheet.cell(row=2, column=1)
	cellDate.alignment = alignRight

	# Model
	sheet.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
	setCellTitleMerge(1, 'Model')

	# Hostname
	sheet.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
	setCellTitleMerge(2, 'Hostname')

	# Version
	sheet.merge_cells(start_row=3, start_column=3, end_row=4, end_column=3)
	setCellTitleMerge(3, 'Version')

	# Uptime
	sheet.merge_cells(start_row=3, start_column=4, end_row=4, end_column=4)
	setCellTitleMerge(4, 'Uptime')

	# CPU
	sheet.merge_cells(start_row=3, start_column=5, end_row=4, end_column=5)
	setCellTitleMerge(5, 'CPU')

	# Memory
	sheet.merge_cells(start_row=3, start_column=6, end_row=3, end_column=7)
	cellMemory = sheet.cell(row=3, column=6, value='Memory')
	setCellStyleTitle(cellMemory)
	cellMemory = sheet.cell(row=3, column=7)
	setCellStyleTitle(cellMemory)
	cellMemoryT = sheet.cell(row=4, column=6, value='Total')
	setCellStyleTitle(cellMemoryT)
	cellMemoryU = sheet.cell(row=4, column=7, value='Used')
	setCellStyleTitle(cellMemoryU)

	# Flash
	sheet.merge_cells(start_row=3, start_column=8, end_row=3, end_column=9)
	cellFlash = sheet.cell(row=3, column=8, value='Flash')
	setCellStyleTitle(cellFlash)
	cellFlash = sheet.cell(row=3, column=9)
	setCellStyleTitle(cellFlash)
	cellFlashT = sheet.cell(row=4, column=8, value='Total')
	setCellStyleTitle(cellFlashT)
	cellFlashF = sheet.cell(row=4, column=9, value='Free')
	setCellStyleTitle(cellFlashF)

	# Temperature
	sheet.merge_cells(start_row=3, start_column=10, end_row=4, end_column=10)
	setCellTitleMerge(10, 'Temp(℃)')

	# Power
	sheet.merge_cells(start_row=3, start_column=11, end_row=4, end_column=11)
	setCellTitleMerge(11, 'Power')

	# Fantray
	sheet.merge_cells(start_row=3, start_column=12, end_row=4, end_column=12)
	setCellTitleMerge(12, 'Fan')

### Set Excel cell Style - alignment Center, border Thin
def setCellStyleDefault(cell):
	cell.alignment = alignCenter
	cell.border = borderThin

### Set Excel cell Style - alignment Center, border Thin, font Bold, color
def setCellStyleTitle(cell):
	cell.alignment = alignCenter
	cell.font = fontBold
	cell.border = borderThin
	cell.fill = fillColor

### Set Excel cell Title and Row Merge
def setCellTitleMerge(column, title):
	sheet.merge_cells(start_row=3, start_column=column, end_row=4, end_column=column)
	cell = sheet.cell(row=3, column=column, value=title)
	setCellStyleTitle(cell)
	cell = sheet.cell(row=4, column=column)
	setCellStyleTitle(cell)


### Main 
global savePath # information excel file path

selectFile = getExcel()
infoFrame = getInformation(selectFile)
savePath = os.path.dirname(os.path.abspath(selectFile)) + "\\"

resultExcel = Workbook()
sheet = resultExcel.active
prepareExcel(sheet)

for index in range(infoFrame.shape[0]):

	# SSH2 connection
	host = infoFrame['host'][index]
	port = infoFrame['port'][index]
	user = infoFrame['username'][index]
	pswd = infoFrame['password'][index]

	connectSSH2(host, port, user, pswd)

	# Connection Succeed
	if crt.Session.Connected:

		crt.Screen.Send("en\r")
		crt.Screen.Send(pswd + "\r")
		crt.Screen.Send("terminal length 0\r")
		
		crt.Screen.Clear()
		crt.Screen.WaitForCursor(1)

		excuteCommand(sheet, index)

		crt.Session.Disconnect()

# Add Date
today = datetime.now()
strDate = 'Date : ' + today.strftime('%Y-%m-%d %H:%M:%S')
cellDate = sheet.cell(row=2, column=1, value=strDate)

# Save Result Excel
try:
	saveFile = crt.Dialog.FileSaveDialog(title="Save As", 
									filter="Excel Files (*.xlsx)|*.xlsx||", 
									defaultFilename="report.xlsx")
except:
	saveFile = savePath + "report_" + today.strftime('%Y%m%d%H%M%S') + ".xlsx"
finally:

	if saveFile is None or saveFile == '':
		saveFile = savePath + "report_" + today.strftime('%Y%m%d%H%M%S') + ".xlsx"
	elif saveFile.find(".") > -1:
		saveFile = saveFile[0:saveFile.find(".")] + ".xlsx"
	else :
		saveFile = saveFile + ".xlsx"

	resultExcel.save(saveFile)
	crt.Dialog.MessageBox("Complete writing Report", "Complete", ICON_INFO)
